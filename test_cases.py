import openpyxl
from openpyxl.styles import Alignment
import os

def create_test_case_1():
    """测试用例1：发货清单"""
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'Sheet1'
    
    # 添加测试数据
    test_data = [
        ['NO.', 'ITEM名称', '规格', '数量', '单位', '备注', '外包装箱号', '经办'],
        ['1', '土壤传感器（DGL定制）', 'DG-214-4G-P2', 20, '台', '标准盒装：主机+探头+背装卡架', 'A箱（1-20号盒）', '刘大鹏'],
        ['2', '振动传感器（DLG定制）', 'PG-T920-4G-P2', 2, '台', '标准盒装：主机+探头', 'A箱-（21-22号盒）', '刘大鹏'],
        ['3', '太阳能储能电池', None, 3, '台', '其中：\n1台世电上海发货（B箱）；\n2台原厂山东发货；（C箱 | D箱）；', 'B | C | D箱', '上海-刘大鹏\n山东-张曼'],
        ['4', '太阳能安装支柱', None, 3, '套', '3套全部原厂山东发货；（E | F | G箱）；', '（E | F | G箱）', '张曼'],
        ['5', '电源线圈', '5米/圈', 22, '套', '需20套，备2套；', 'A箱-23号盒', '刘大鹏'],
    ]
    
    for row_idx, row_data in enumerate(test_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.value = value
    
    return wb

def create_test_case_2():
    """测试用例2：产品列表"""
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'Sheet1'
    
    # 添加测试数据
    test_data = [
        ['产品编号', '产品名称', '产品描述', '单价（USD）', '库存数量', '状态'],
        ['P001', '压力传感器', '用于工业压力监测', 150.00, 100, '有库存'],
        ['P002', '温度传感器', '用于温度监测', 120.00, 50, '有库存'],
        ['P003', '液位传感器', '用于液位监测', 180.00, 0, '缺货'],
        ['P004', '流量传感器', '用于流量监测', 200.00, 25, '有库存'],
    ]
    
    for row_idx, row_data in enumerate(test_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.value = value
    
    return wb

def create_test_case_3():
    """测试用例3：订单信息"""
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'Sheet1'
    
    # 添加测试数据
    test_data = [
        ['订单编号', '客户名称', '产品名称', '数量', '单价（USD）', '总价（USD）', '订单状态', '交货日期'],
        ['ORD-2024-001', 'ABC公司', '压力传感器', 10, 150.00, 1500.00, '已确认', '2024-03-15'],
        ['ORD-2024-002', 'XYZ公司', '温度传感器', 5, 120.00, 600.00, '待确认', '2024-03-20'],
        ['ORD-2024-003', 'DEF公司', '液位传感器', 8, 180.00, 1440.00, '已发货', '2024-03-10'],
        ['ORD-2024-004', 'GHI公司', '流量传感器', 3, 200.00, 600.00, '已取消', '2024-03-25'],
    ]
    
    for row_idx, row_data in enumerate(test_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.value = value
    
    return wb

def save_test_case(wb, filename):
    """保存测试用例"""
    wb.save(filename)
    print(f'创建测试文件: {filename}')
    return filename

def run_test_cases():
    """运行所有测试用例"""
    import sys
    import os
    sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
    from scripts.bilingual_excel_sync import BilingualExcelSync
    
    # 创建测试用例
    test_cases = [
        ('测试用例1_发货清单.xlsx', create_test_case_1()),
        ('测试用例2_产品列表.xlsx', create_test_case_2()),
        ('测试用例3_订单信息.xlsx', create_test_case_3()),
    ]
    
    # 创建同步器实例
    syncer = BilingualExcelSync()
    
    # 加载翻译字典
    translations = {
        # 通用字段
        'NO.': 'NO.',
        'ITEM名称': 'Item Name',
        '规格': 'Specification',
        '数量': 'Quantity',
        '单位': 'Unit',
        '备注': 'Remarks',
        '外包装箱号': 'Packing Box No.',
        '经办': 'Handler',
        '产品编号': 'Product Code',
        '产品名称': 'Product Name',
        '产品描述': 'Product Description',
        '单价（USD）': 'Unit Price (USD)',
        '库存数量': 'Stock Quantity',
        '状态': 'Status',
        '订单编号': 'Order Number',
        '客户名称': 'Customer Name',
        '总价（USD）': 'Total Price (USD)',
        '订单状态': 'Order Status',
        '交货日期': 'Delivery Date',
        
        # 产品名称
        '土壤传感器（DGL定制）': 'Soil Sensor (DGL Customized)',
        '振动传感器（DLG定制）': 'Vibration Sensor (DLG Customized)',
        '太阳能储能电池': 'Solar Energy Storage Battery',
        '太阳能安装支柱': 'Solar Installation Support Pole',
        '电源线圈': 'Power Coil',
        '压力传感器': 'Pressure Sensor',
        '温度传感器': 'Temperature Sensor',
        '液位传感器': 'Level Sensor',
        '流量传感器': 'Flow Sensor',
        
        # 单位
        '台': 'Unit/Set',
        '套': 'Set',
        '个': 'Piece',
        
        # 状态
        '有库存': 'In Stock',
        '缺货': 'Out of Stock',
        '已确认': 'Confirmed',
        '待确认': 'Pending Confirmation',
        '已发货': 'Shipped',
        '已取消': 'Cancelled',
        
        # 备注
        '标准盒装：主机+探头+背装卡架': 'Standard Box: Main Unit + Probe + Back-Mount Card Rack',
        '标准盒装：主机+探头': 'Standard Box: Main Unit + Probe',
        '其中：\n1台世电上海发货（B箱）；\n2台原厂山东发货；（C箱 | D箱）；': 'Including:\n1 unit shipped from WEST Shanghai (Box B);\n2 units shipped from original factory Shandong; (Box C | D);',
        '3套全部原厂山东发货；（E | F | G箱）；': 'All 3 sets shipped from original factory Shandong; (Box E | F | G);',
        '需20套，备2套；': 'Need 20 sets, 2 sets spare;',
        '用于工业压力监测': 'For industrial pressure monitoring',
        '用于温度监测': 'For temperature monitoring',
        '用于液位监测': 'For level monitoring',
        '用于流量监测': 'For flow monitoring',
        
        # 经办人
        '刘大鹏': 'Liu Dapeng',
        '张曼': 'Zhang Man',
        '上海-刘大鹏\n山东-张曼': 'Shanghai-Liu Dapeng\nShandong-Zhang Man',
        
        # 客户名称
        'ABC公司': 'ABC Company',
        'XYZ公司': 'XYZ Company',
        'DEF公司': 'DEF Company',
        'GHI公司': 'GHI Company',
    }
    
    syncer.load_translations(translations)
    
    # 运行测试用例
    for filename, wb in test_cases:
        # 保存测试文件
        save_test_case(wb, filename)
        
        # 执行转换
        output_filename = filename.replace('.xlsx', '_中英双语版.xlsx')
        syncer.convert(filename, output_filename)
        
        # 清理测试文件
        if os.path.exists(filename):
            os.remove(filename)
            print(f'清理测试文件: {filename}')
    
    print('\n所有测试用例运行完成！')

if __name__ == '__main__':
    run_test_cases()