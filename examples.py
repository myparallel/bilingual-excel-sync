import openpyxl
from openpyxl.styles import Alignment
import os

def example_1_basic_usage():
    """示例1：基本使用"""
    print("=== 示例1：基本使用 ===")
    
    import sys
    import os
    sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
    from scripts.bilingual_excel_sync import BilingualExcelSync
    
    # 创建测试文件
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'Sheet1'
    
    # 添加测试数据
    test_data = [
        ['项目名称', '数量', '状态'],
        ['传感器', 10, '已付'],
        ['变送器', 5, '待付'],
    ]
    
    for row_idx, row_data in enumerate(test_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.value = value
    
    # 保存测试文件
    test_file = '示例1_测试.xlsx'
    wb.save(test_file)
    
    # 创建同步器实例
    syncer = BilingualExcelSync()
    
    # 加载翻译字典
    translations = {
        '项目名称': 'Project Name',
        '数量': 'Quantity',
        '状态': 'Status',
        '传感器': 'Sensor',
        '变送器': 'Transmitter',
        '已付': 'Paid',
        '待付': 'Pending',
    }
    
    syncer.load_translations(translations)
    
    # 执行转换
    output_file = '示例1_双语版.xlsx'
    syncer.convert(test_file, output_file)
    
    # 清理测试文件
    if os.path.exists(test_file):
        os.remove(test_file)
    
    print(f"输出文件: {output_file}")
    print()

def example_2_custom_translations():
    """示例2：自定义翻译"""
    print("=== 示例2：自定义翻译 ===")
    
    import sys
    import os
    sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
    from scripts.bilingual_excel_sync import BilingualExcelSync
    
    # 创建测试文件
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'Sheet1'
    
    # 添加测试数据
    test_data = [
        ['产品名称', '规格', '单价'],
        ['土壤传感器', 'DG-214', 150.00],
        ['振动传感器', 'PG-T920', 200.00],
    ]
    
    for row_idx, row_data in enumerate(test_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.value = value
    
    # 保存测试文件
    test_file = '示例2_测试.xlsx'
    wb.save(test_file)
    
    # 创建同步器实例
    syncer = BilingualExcelSync()
    
    # 加载自定义翻译字典
    custom_translations = {
        '产品名称': 'Product Name',
        '规格': 'Specification',
        '单价': 'Unit Price',
        '土壤传感器': 'Soil Sensor',
        '振动传感器': 'Vibration Sensor',
    }
    
    syncer.load_translations(custom_translations)
    
    # 执行转换
    output_file = '示例2_双语版.xlsx'
    syncer.convert(test_file, output_file)
    
    # 清理测试文件
    if os.path.exists(test_file):
        os.remove(test_file)
    
    print(f"输出文件: {output_file}")
    print()

def example_3_sync_update():
    """示例3：同步更新"""
    print("=== 示例3：同步更新 ===")
    
    import sys
    import os
    sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
    from scripts.bilingual_excel_sync import BilingualExcelSync
    
    # 创建测试文件
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'Sheet1'
    
    # 添加测试数据
    test_data = [
        ['项目名称', '状态'],
        ['传感器', '已付'],
        ['变送器', '待付'],
    ]
    
    for row_idx, row_data in enumerate(test_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.value = value
    
    # 保存测试文件
    test_file = '示例3_测试.xlsx'
    wb.save(test_file)
    
    # 创建同步器实例
    syncer = BilingualExcelSync()
    
    # 加载翻译字典
    translations = {
        '项目名称': 'Project Name',
        '状态': 'Status',
        '传感器': 'Sensor',
        '变送器': 'Transmitter',
        '已付': 'Paid',
        '待付': 'Pending',
    }
    
    syncer.load_translations(translations)
    
    # 执行转换
    output_file = '示例3_双语版.xlsx'
    syncer.convert(test_file, output_file)
    
    # 模拟更新SHEET1
    wb = openpyxl.load_workbook(output_file)
    sheet1 = wb[wb.sheetnames[0]]
    
    # 更新第2行第2列的值
    cell = sheet1.cell(row=2, column=2)
    cell.value = '传感器\nSensor'  # 确保是双语格式
    
    # 保存更新
    wb.save(output_file)
    
    # 同步更新SHEET2
    syncer.load_workbook(output_file)
    syncer.sync_english_sheet()
    syncer.save(output_file)
    
    # 清理测试文件
    if os.path.exists(test_file):
        os.remove(test_file)
    
    print(f"输出文件: {output_file}")
    print("已同步更新SHEET2")
    print()

def example_4_verify_translation():
    """示例4：验证翻译"""
    print("=== 示例4：验证翻译 ===")
    
    import sys
    import os
    sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
    from scripts.bilingual_excel_sync import BilingualExcelSync
    
    # 创建测试文件
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'Sheet1'
    
    # 添加测试数据
    test_data = [
        ['项目名称', '状态'],
        ['传感器', '已付'],
        ['变送器', '待付'],
        ['未知项目', '未知状态'],
    ]
    
    for row_idx, row_data in enumerate(test_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.value = value
    
    # 保存测试文件
    test_file = '示例4_测试.xlsx'
    wb.save(test_file)
    
    # 创建同步器实例
    syncer = BilingualExcelSync()
    
    # 加载翻译字典（故意不完整，用于演示验证功能）
    translations = {
        '项目名称': 'Project Name',
        '状态': 'Status',
        '传感器': 'Sensor',
        '变送器': 'Transmitter',
        '已付': 'Paid',
        '待付': 'Pending',
    }
    
    syncer.load_translations(translations)
    
    # 执行转换
    output_file = '示例4_双语版.xlsx'
    syncer.convert(test_file, output_file)
    
    # 验证翻译
    verification_file = '示例4_验证结果.txt'
    syncer.verify(output_file, verification_file)
    
    # 清理测试文件
    if os.path.exists(test_file):
        os.remove(test_file)
    
    print(f"输出文件: {output_file}")
    print(f"验证结果: {verification_file}")
    print()

def run_all_examples():
    """运行所有示例"""
    print("运行所有示例...")
    print()
    
    example_1_basic_usage()
    example_2_custom_translations()
    example_3_sync_update()
    example_4_verify_translation()
    
    print("所有示例运行完成！")

if __name__ == '__main__':
    run_all_examples()