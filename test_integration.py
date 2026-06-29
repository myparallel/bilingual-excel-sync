import openpyxl
from openpyxl.styles import Alignment
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from scripts.bilingual_excel_sync import BilingualExcelSync

def test_integration():
    """集成测试"""
    print("=== 集成测试开始 ===")
    
    # 1. 创建测试文件
    print("1. 创建测试文件...")
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'Sheet1'
    
    test_data = [
        ['NO.', '项目名称', '规格', '数量', '单位', '状态', '备注'],
        ['1', '土壤传感器', 'DG-214', 20, '台', '已确认', '标准盒装'],
        ['2', '振动传感器', 'PG-T920', 2, '台', '待确认', '特殊定制'],
        ['3', '太阳能储能电池', None, 3, '台', '已发货', '上海发货'],
        ['4', '电源线圈', '5米/圈', 22, '套', '有库存', '热销产品'],
    ]
    
    for row_idx, row_data in enumerate(test_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.value = value
    
    test_file = 'integration_test.xlsx'
    wb.save(test_file)
    print(f"   创建测试文件: {test_file}")
    
    # 2. 创建同步器实例
    print("2. 创建同步器实例...")
    syncer = BilingualExcelSync()
    
    # 3. 加载翻译字典
    print("3. 加载翻译字典...")
    translations = {
        # 列标题
        'NO.': 'NO.',
        '项目名称': 'Project Name',
        '规格': 'Specification',
        '数量': 'Quantity',
        '单位': 'Unit',
        '状态': 'Status',
        '备注': 'Remarks',
        
        # 项目名称
        '土壤传感器': 'Soil Sensor',
        '振动传感器': 'Vibration Sensor',
        '太阳能储能电池': 'Solar Energy Storage Battery',
        '电源线圈': 'Power Coil',
        
        # 单位
        '台': 'Unit/Set',
        '套': 'Set',
        
        # 状态
        '已确认': 'Confirmed',
        '待确认': 'Pending Confirmation',
        '已发货': 'Shipped',
        '有库存': 'In Stock',
        
        # 备注
        '标准盒装': 'Standard Box',
        '特殊定制': 'Special Customization',
        '上海发货': 'Shanghai Shipment',
        '热销产品': 'Hot Selling Product',
    }
    
    syncer.load_translations(translations)
    print(f"   加载翻译字典: {len(translations)} 个条目")
    
    # 4. 执行转换
    print("4. 执行转换...")
    output_file = 'integration_test_中英双语版.xlsx'
    syncer.convert(test_file, output_file)
    print(f"   输出文件: {output_file}")
    
    # 5. 验证结果
    print("5. 验证结果...")
    wb = openpyxl.load_workbook(output_file)
    
    # 检查工作表数量
    assert len(wb.sheetnames) == 2, f"期望2个工作表，实际{len(wb.sheetnames)}个"
    print(f"   工作表数量: {len(wb.sheetnames)} OK")
    
    # 检查工作表名称
    assert 'Sheet1 双语版 Bilingual' in wb.sheetnames, "缺少双语版工作表"
    assert 'Sheet2 English Version' in wb.sheetnames, "缺少英文版工作表"
    print(f"   工作表名称: {wb.sheetnames} OK")
    
    # 检查SHEET1内容
    sheet1 = wb['Sheet1 双语版 Bilingual']
    print(f"   SHEET1行数: {sheet1.max_row} OK")
    print(f"   SHEET1列数: {sheet1.max_column} OK")
    
    # 检查SHEET2内容
    sheet2 = wb['Sheet2 English Version']
    print(f"   SHEET2行数: {sheet2.max_row} OK")
    print(f"   SHEET2列数: {sheet2.max_column} OK")
    
    # 检查翻译完整性
    print("6. 检查翻译完整性...")
    for row_idx, row in enumerate(sheet1.iter_rows(values_only=True), 1):
        for col_idx, cell in enumerate(row, 1):
            if cell and isinstance(cell, str):
                if '\n' in cell:
                    # 双语格式
                    chinese, english = cell.split('\n', 1)
                    print(f"   行{row_idx}, 列{col_idx}: {chinese} -> {english} OK")
                else:
                    # 检查是否是数字或公式
                    if not cell.replace('.', '').isdigit() and not cell.startswith('='):
                        print(f"   行{row_idx}, 列{col_idx}: {cell} (未翻译)")
    
    # 7. 检查SHEET2是否只包含英文
    print("7. 检查SHEET2英文内容...")
    for row_idx, row in enumerate(sheet2.iter_rows(values_only=True), 1):
        for col_idx, cell in enumerate(row, 1):
            if cell and isinstance(cell, str):
                # 检查是否包含中文字符
                import re
                chinese_pattern = re.compile(r'[\u4e00-\u9fff]')
                if chinese_pattern.search(cell):
                    print(f"   警告: SHEET2包含中文字符: 行{row_idx}, 列{col_idx}: {cell}")
    
    # 8. 清理测试文件
    print("8. 清理测试文件...")
    if os.path.exists(test_file):
        os.remove(test_file)
        print(f"   删除测试文件: {test_file}")
    
    print("=== 集成测试完成 ===")
    print(f"输出文件: {output_file}")
    print("请打开文件查看效果！")

if __name__ == '__main__':
    test_integration()