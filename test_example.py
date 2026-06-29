import openpyxl
from openpyxl.styles import Alignment
import os

def create_test_excel():
    """创建测试用的Excel文件"""
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'Sheet1'
    
    # 添加测试数据
    test_data = [
        ['NO.', '项目名称', '规格', '数量', '单位', '备注'],
        ['1', '土壤传感器', 'DG-214-4G-P2', 20, '台', '标准盒装'],
        ['2', '振动传感器', 'PG-T920-4G-P2', 2, '台', '标准盒装'],
        ['3', '太阳能储能电池', None, 3, '台', '上海发货'],
        ['4', '太阳能安装支柱', None, 3, '套', '山东发货'],
    ]
    
    for row_idx, row_data in enumerate(test_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.value = value
    
    # 保存测试文件
    test_file = '测试表格.xlsx'
    wb.save(test_file)
    print(f'创建测试文件: {test_file}')
    return test_file

def test_bilingual_sync():
    """测试双语同步功能"""
    import sys
    import os
    sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
    from scripts.bilingual_excel_sync import BilingualExcelSync
    
    # 创建测试文件
    test_file = create_test_excel()
    
    # 创建同步器实例
    syncer = BilingualExcelSync()
    
    # 加载翻译字典
    translations = {
        '项目名称': 'Project Name',
        '规格': 'Specification',
        '数量': 'Quantity',
        '单位': 'Unit',
        '备注': 'Remarks',
        '土壤传感器': 'Soil Sensor',
        '振动传感器': 'Vibration Sensor',
        '太阳能储能电池': 'Solar Energy Storage Battery',
        '太阳能安装支柱': 'Solar Installation Support Pole',
        '台': 'Unit/Set',
        '套': 'Set',
        '标准盒装': 'Standard Box',
        '上海发货': 'Shanghai Shipment',
        '山东发货': 'Shandong Shipment',
    }
    
    syncer.load_translations(translations)
    
    # 执行转换
    output_file = '测试表格_中英双语版.xlsx'
    syncer.convert(test_file, output_file)
    
    # 验证结果
    wb = openpyxl.load_workbook(output_file)
    
    print(f'\n=== 验证结果 ===')
    print(f'工作表列表: {wb.sheetnames}')
    
    # 检查SHEET1
    sheet1 = wb[wb.sheetnames[0]]
    print(f'\nSHEET1内容:')
    for row_idx, row in enumerate(sheet1.iter_rows(values_only=True), 1):
        if row_idx <= 3:  # 只显示前3行
            print(f'行{row_idx}: {row}')
    
    # 检查SHEET2
    if len(wb.sheetnames) > 1:
        sheet2 = wb[wb.sheetnames[1]]
        print(f'\nSHEET2内容:')
        for row_idx, row in enumerate(sheet2.iter_rows(values_only=True), 1):
            if row_idx <= 3:  # 只显示前3行
                print(f'行{row_idx}: {row}')
    
    # 清理测试文件
    if os.path.exists(test_file):
        os.remove(test_file)
        print(f'\n清理测试文件: {test_file}')
    
    print(f'\n测试完成！输出文件: {output_file}')

if __name__ == '__main__':
    test_bilingual_sync()