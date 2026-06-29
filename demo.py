import openpyxl
from openpyxl.styles import Alignment
import os

def create_demo_excel():
    """创建演示用的Excel文件"""
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'Sheet1'
    
    # 添加演示数据
    demo_data = [
        ['NO.', '产品名称', '产品描述', '规格型号', '数量', '单位', '单价（USD）', '总价（USD）', '状态', '备注'],
        ['1', '压力传感器', '用于工业压力监测', 'PS-200', 10, '个', 150.00, 1500.00, '已确认', '标准包装'],
        ['2', '温度传感器', '用于温度监测', 'TS-100', 5, '个', 120.00, 600.00, '待确认', '含配件'],
        ['3', '液位传感器', '用于液位监测', 'LS-300', 8, '个', 180.00, 1440.00, '已发货', '特殊定制'],
        ['4', '流量传感器', '用于流量监测', 'FS-400', 3, '个', 200.00, 600.00, '已取消', '客户取消'],
        ['5', '湿度传感器', '用于湿度监测', 'HS-500', 15, '个', 90.00, 1350.00, '有库存', '热销产品'],
    ]
    
    for row_idx, row_data in enumerate(demo_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.value = value
    
    # 保存演示文件
    demo_file = '演示表格.xlsx'
    wb.save(demo_file)
    print(f'创建演示文件: {demo_file}')
    return demo_file

def run_demo():
    """运行演示"""
    import sys
    import os
    sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
    from scripts.bilingual_excel_sync import BilingualExcelSync
    
    # 创建演示文件
    demo_file = create_demo_excel()
    
    # 创建同步器实例
    syncer = BilingualExcelSync()
    
    # 加载翻译字典
    translations = {
        # 列标题
        'NO.': 'NO.',
        '产品名称': 'Product Name',
        '产品描述': 'Product Description',
        '规格型号': 'Model Number',
        '数量': 'Quantity',
        '单位': 'Unit',
        '单价（USD）': 'Unit Price (USD)',
        '总价（USD）': 'Total Price (USD)',
        '状态': 'Status',
        '备注': 'Remarks',
        
        # 产品名称
        '压力传感器': 'Pressure Sensor',
        '温度传感器': 'Temperature Sensor',
        '液位传感器': 'Level Sensor',
        '流量传感器': 'Flow Sensor',
        '湿度传感器': 'Humidity Sensor',
        
        # 产品描述
        '用于工业压力监测': 'For industrial pressure monitoring',
        '用于温度监测': 'For temperature monitoring',
        '用于液位监测': 'For level monitoring',
        '用于流量监测': 'For flow monitoring',
        '用于湿度监测': 'For humidity monitoring',
        
        # 单位
        '个': 'Piece',
        
        # 状态
        '已确认': 'Confirmed',
        '待确认': 'Pending Confirmation',
        '已发货': 'Shipped',
        '已取消': 'Cancelled',
        '有库存': 'In Stock',
        
        # 备注
        '标准包装': 'Standard Packaging',
        '含配件': 'With Accessories',
        '特殊定制': 'Special Customization',
        '客户取消': 'Customer Cancelled',
        '热销产品': 'Hot Selling Product',
    }
    
    syncer.load_translations(translations)
    
    # 执行转换
    output_file = '演示表格_中英双语版.xlsx'
    syncer.convert(demo_file, output_file)
    
    # 验证结果
    wb = openpyxl.load_workbook(output_file)
    
    print(f'\n=== 演示结果 ===')
    print(f'工作表列表: {wb.sheetnames}')
    
    # 检查SHEET1
    sheet1 = wb[wb.sheetnames[0]]
    print(f'\nSHEET1（双语版）内容:')
    for row_idx, row in enumerate(sheet1.iter_rows(values_only=True), 1):
        if row_idx <= 4:  # 只显示前4行
            print(f'行{row_idx}: {row}')
    
    # 检查SHEET2
    if len(wb.sheetnames) > 1:
        sheet2 = wb[wb.sheetnames[1]]
        print(f'\nSHEET2（纯英文版）内容:')
        for row_idx, row in enumerate(sheet2.iter_rows(values_only=True), 1):
            if row_idx <= 4:  # 只显示前4行
                print(f'行{row_idx}: {row}')
    
    # 清理演示文件
    if os.path.exists(demo_file):
        os.remove(demo_file)
        print(f'\n清理演示文件: {demo_file}')
    
    print(f'\n演示完成！输出文件: {output_file}')
    print(f'\n请打开 {output_file} 查看效果：')
    print(f'1. SHEET1：中英双语版（中文在上，英文在下）')
    print(f'2. SHEET2：纯英文版（只包含英文内容）')

if __name__ == '__main__':
    run_demo()