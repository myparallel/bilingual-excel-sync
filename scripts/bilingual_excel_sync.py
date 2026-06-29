import openpyxl
from openpyxl.styles import Alignment, Font, Border, PatternFill
from copy import copy
import re
import sys
import os

class BilingualExcelSync:
    def __init__(self):
        self.wb = None
        self.translations = {}
        self.sheet_translations = {}
        
    def load_workbook(self, file_path):
        """加载Excel文件"""
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"文件不存在: {file_path}")
        self.wb = openpyxl.load_workbook(file_path)
        return self.wb
    
    def load_translations(self, translations):
        """批量加载翻译字典"""
        self.translations.update(translations)
    
    def add_sheet_translation(self, chinese, english):
        """添加工作表名称翻译"""
        self.sheet_translations[chinese] = english
    
    def contains_chinese(self, text):
        """检查文本是否包含中文字符"""
        chinese_pattern = re.compile(r'[\u4e00-\u9fff]')
        return bool(chinese_pattern.search(text))
    
    def extract_english(self, text):
        """从双语文本中提取英文部分"""
        if text is None:
            return None
        if isinstance(text, (int, float)):
            return text
        text_str = str(text)
        if '\n' in text_str:
            lines = text_str.split('\n')
            # 找到第一个不包含中文字符的行
            for i, line in enumerate(lines):
                if not self.contains_chinese(line):
                    # 返回从这一行开始的所有行
                    return '\n'.join(lines[i:])
        return text_str
    
    def analyze(self, input_file, output_file):
        """分析文件内容，输出所有唯一字符串到文件"""
        if not self.wb:
            self.load_workbook(input_file)
        
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(f'文件: {input_file}\n')
            f.write(f'工作表: {self.wb.sheetnames}\n\n')
            
            for sheet_name in self.wb.sheetnames:
                sheet = self.wb[sheet_name]
                f.write(f'=== 工作表: {sheet_name} ===\n')
                f.write(f'行数: {sheet.max_row}, 列数: {sheet.max_column}\n\n')
                
                # 收集唯一值
                unique_values = set()
                for row in sheet.iter_rows(values_only=True):
                    for cell in row:
                        if cell and isinstance(cell, str):
                            unique_values.add(cell)
                
                f.write('所有唯一字符串值:\n')
                for value in sorted(unique_values):
                    f.write(f'  "{value}" (len={len(value)})\n')
                
                f.write('\n所有行数据:\n')
                for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
                    f.write(f'Row {row_idx}: {row}\n')
        
        return output_file
    
    def convert_to_bilingual(self):
        """将SHEET1转换为双语格式"""
        if not self.wb:
            raise ValueError("请先加载工作簿")
        
        # 获取第一个工作表
        sheet_name = self.wb.sheetnames[0]
        sheet = self.wb[sheet_name]
        
        # 更新工作表名称
        new_sheet_name = self.sheet_translations.get(sheet_name, f'{sheet_name} 双语版 Bilingual')
        sheet.title = new_sheet_name
        
        # 处理单元格
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    if cell.value in self.translations:
                        # 添加英文翻译到中文下方
                        cell.value = cell.value + '\n' + self.translations[cell.value]
                        cell.alignment = Alignment(wrap_text=True)
        
        return self.wb
    
    def sync_english_sheet(self):
        """同步生成纯英文SHEET2"""
        if not self.wb:
            raise ValueError("请先加载工作簿")
        
        # 获取第一个工作表（双语版）
        sheet1_name = self.wb.sheetnames[0]
        sheet1 = self.wb[sheet1_name]
        
        # 删除旧的SHEET2（如果存在）
        if 'Sheet2 English Version' in self.wb.sheetnames:
            del self.wb['Sheet2 English Version']
        
        # 创建新的SHEET2
        sheet2 = self.wb.create_sheet('Sheet2 English Version')
        
        # 复制SHEET1内容到SHEET2，提取英文部分
        for row_idx, row in enumerate(sheet1.iter_rows(), 1):
            for col_idx, cell in enumerate(row, 1):
                new_cell = sheet2.cell(row=row_idx, column=col_idx)
                
                if cell.value is not None:
                    english_value = self.extract_english(cell.value)
                    new_cell.value = english_value
                
                # 复制样式
                if cell.font:
                    new_cell.font = copy(cell.font)
                if cell.alignment:
                    new_cell.alignment = copy(cell.alignment)
                if cell.border:
                    new_cell.border = copy(cell.border)
                if cell.fill:
                    new_cell.fill = copy(cell.fill)
        
        # 复制列宽
        for col_letter, col_dim in sheet1.column_dimensions.items():
            sheet2.column_dimensions[col_letter].width = col_dim.width
        
        # 复制行高
        for row_num, row_dim in sheet1.row_dimensions.items():
            sheet2.row_dimensions[row_num].height = row_dim.height
        
        # 设置页面为A4横向
        sheet2.page_setup.paperSize = sheet2.PAPERSIZE_A4
        sheet2.page_setup.orientation = 'landscape'
        
        return self.wb
    
    def adjust_layout(self):
        """调整布局为A4横向版面"""
        if not self.wb:
            raise ValueError("请先加载工作簿")
        
        for sheet_name in self.wb.sheetnames:
            sheet = self.wb[sheet_name]
            
            # 设置A4横向版面
            sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
            sheet.page_setup.orientation = 'landscape'
            
            # 调整行高以适应双语内容
            for row_num in range(1, sheet.max_row + 1):
                max_lines = 1
                for col_num in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=row_num, column=col_num)
                    if cell.value and isinstance(cell.value, str):
                        lines = cell.value.count('\n') + 1
                        max_lines = max(max_lines, lines)
                
                if max_lines > 1:
                    sheet.row_dimensions[row_num].height = max(15 * max_lines, 30)
        
        return self.wb
    
    def convert(self, input_file, output_file=None):
        """执行完整转换流程"""
        if not output_file:
            name, ext = os.path.splitext(input_file)
            output_file = f"{name}_中英双语版{ext}"
        
        # 加载工作簿
        print(f"加载文件: {input_file}")
        if not self.wb:
            self.load_workbook(input_file)
        
        # 转换为双语格式
        print("转换为双语格式...")
        self.convert_to_bilingual()
        
        # 同步生成英文表
        print("同步生成英文表...")
        self.sync_english_sheet()
        
        # 调整布局
        print("调整布局...")
        self.adjust_layout()
        
        # 保存文件
        print(f"保存文件: {output_file}")
        self.save(output_file)
        
        print("转换完成！")
        return output_file
    
    def verify(self, bilingual_file, output_file):
        """验证转换结果，检查未翻译的字段"""
        wb = openpyxl.load_workbook(bilingual_file)
        
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(f'验证文件: {bilingual_file}\n\n')
            
            total_cells = 0
            translated_cells = 0
            untranslated = []
            
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                f.write(f'=== 工作表: {sheet_name} ===\n')
                
                for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
                    for col_idx, cell in enumerate(row, 1):
                        if cell and isinstance(cell, str):
                            total_cells += 1
                            if '\n' in cell:
                                translated_cells += 1
                            else:
                                # 检查是否是公式或数字
                                if not cell.startswith('=') and not cell.replace('.', '').isdigit():
                                    untranslated.append(f'Row {row_idx}, Col {col_idx}: "{cell}"')
                
                # 输出所有行
                f.write('\n所有行数据:\n')
                for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
                    f.write(f'Row {row_idx}: {row}\n')
            
            f.write(f'\n=== 统计 ===\n')
            f.write(f'总单元格数: {total_cells}\n')
            f.write(f'已翻译单元格数: {translated_cells}\n')
            f.write(f'未翻译单元格数: {len(untranslated)}\n')
            
            if untranslated:
                f.write(f'\n未翻译的字段:\n')
                for item in untranslated:
                    f.write(f'  {item}\n')
        
        return output_file
    
    def save(self, output_path):
        """保存双语版Excel文件"""
        if not self.wb:
            raise ValueError("请先加载工作簿")
        self.wb.save(output_path)
        return output_path


# 通用翻译字典
COMMON_TRANSLATIONS = {
    # 列标题
    'NO.': 'NO.',
    '项目名称': 'Project Name',
    '项目描述': 'Project Description',
    '金额（USD)': 'Amount (USD)',
    '状态': 'Status',
    '传感器类型': 'Sensor Type',
    '传感器型号': 'Sensor Model',
    '基础型号': 'Base Model',
    '传感类型': 'Sensor Type',
    
    # 状态值
    '已付': 'Paid',
    '待付': 'Pending',
    '未付': 'Unpaid',
    '已签约': 'Signed',
    '待签约': 'To Be Signed',
    '谈判中': 'In Negotiation',
    '待谈判': 'To Be Negotiated',
    '未签约': 'Not Signed',
    
    # 汇总行
    '小计：': 'Subtotal:',
    '总计：': 'Total:',
    
    # 日期格式
    '（预估）': '(Estimated)',
    '预计': 'Expected',
}

# 外贸常用术语
TRADE_TRANSLATIONS = {
    # 贸易术语
    '发货清单': 'Shipping List',
    '装箱单': 'Packing List',
    '发票': 'Invoice',
    '合同': 'Contract',
    '订单': 'Order',
    '样品': 'Sample',
    
    # 产品相关
    '产品名称': 'Product Name',
    '产品描述': 'Product Description',
    '规格': 'Specification',
    '数量': 'Quantity',
    '单位': 'Unit',
    '单价': 'Unit Price',
    '总价': 'Total Price',
    
    # 物流相关
    '发货': 'Shipment',
    '收货': 'Receipt',
    '运输': 'Transport',
    '仓储': 'Warehousing',
    '报关': 'Customs Declaration',
    
    # 付款相关
    '付款方式': 'Payment Method',
    '预付款': 'Advance Payment',
    '尾款': 'Balance Payment',
    '信用证': 'Letter of Credit',
    '电汇': 'Wire Transfer',
}

# 传感器/工业术语
SENSOR_TRANSLATIONS = {
    '压力监测变送器': 'Pressure Monitoring Transmitter',
    '液位监测变送器': 'Level Monitoring Transmitter',
    '微压差变送器': 'Micro Differential Pressure Transmitter',
    '温度监测变送器': 'Temperature Monitoring Transmitter',
    '温湿度监测变送器': 'Temperature & Humidity Monitoring Transmitter',
    'PH监测变送器': 'PH Monitoring Transmitter',
    '电导率监测变送器': 'Conductivity Monitoring Transmitter',
    '溶解氧变送器': 'Dissolved Oxygen Transmitter',
    '浊度变送器': 'Turbidity Transmitter',
    '双通道压力监测变送器': 'Dual-Channel Pressure Monitoring Transmitter',
    '水浸监测变送器': 'Water Immersion Monitoring Transmitter',
    '土壤温湿度电导率变送器': 'Soil Temp/Humidity/Conductivity Transmitter',
    '防爆压力变送器': 'Explosion-Proof Pressure Transmitter',
    '防爆液位变送器': 'Explosion-Proof Level Transmitter',
    '防爆深潜投入式液位变送器': 'Explosion-Proof Submersible Level Transmitter',
    '氨气传感变送器': 'Ammonia Gas Sensor Transmitter',
    '雷达物位变送器': 'Radar Level Transmitter',
    '液位尺传感变送器': 'Level Ruler Sensor Transmitter',
    '二氧化碳传感变送器': 'CO2 Sensor Transmitter',
    '甲醛传感变送器': 'Formaldehyde Sensor Transmitter',
    '多普勒流量计': 'Doppler Flow Meter',
    '甲烷传感变送器': 'Methane Sensor Transmitter',
    '一氧化碳传感变送器': 'CO Sensor Transmitter',
    '氧气传感变送器': 'Oxygen Sensor Transmitter',
    '超声波液位传感变送器': 'Ultrasonic Level Sensor Transmitter',
    '防爆温度变送器': 'Explosion-Proof Temperature Transmitter',
    '盐度传感变送器': 'Salinity Sensor Transmitter',
    '疏水阀传感器': 'Steam Trap Sensor',
    '多功能检测仪': 'Multi-function Detector',
    '三通道压力及甲烷泄露监测传感变送器': '3-Channel Pressure & Methane Leak Monitor Transmitter',
    '电导率变送器': 'Conductivity Transmitter',
    '硫化氢传感变送器': 'H2S Sensor Transmitter',
}


def main():
    """主函数"""
    if len(sys.argv) < 2:
        print("使用方法:")
        print("  分析: python bilingual_excel_sync.py analyze <input_file> [output_file]")
        print("  转换: python bilingual_excel_sync.py convert <input_file> [output_file]")
        print("  验证: python bilingual_excel_sync.py verify <bilingual_file> [output_file]")
        sys.exit(1)
    
    command = sys.argv[1]
    input_file = sys.argv[2]
    output_file = sys.argv[3] if len(sys.argv) > 3 else None
    
    syncer = BilingualExcelSync()
    
    # 加载通用翻译
    syncer.load_translations(COMMON_TRANSLATIONS)
    syncer.load_translations(TRADE_TRANSLATIONS)
    syncer.load_translations(SENSOR_TRANSLATIONS)
    
    if command == 'analyze':
        if not output_file:
            output_file = 'analysis.txt'
        syncer.analyze(input_file, output_file)
        print(f"分析结果已保存到: {output_file}")
    
    elif command == 'convert':
        syncer.convert(input_file, output_file)
    
    elif command == 'verify':
        if not output_file:
            output_file = 'verification.txt'
        syncer.verify(input_file, output_file)
        print(f"验证结果已保存到: {output_file}")
    
    else:
        print(f"未知命令: {command}")
        sys.exit(1)


if __name__ == "__main__":
    main()